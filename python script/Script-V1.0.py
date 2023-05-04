import json
import os
from datetime import datetime

import openpyxl


def validate_path(path):
    if path is None or not os.path.exists(path):
        raise Exception("Path does not exist ")


def read_config():
    config = {}
    with open('config.txt', 'r') as f:
        for line in f:
            line = line.strip()
            if line.startswith('#') or len(line) == 0:
                continue
            if line.startswith('['):
                section = line[1:-1]
                section = section.strip()
                config[section] = {}
            else:
                key, value = line.split('=')
                config[section][key.strip()] = value.strip()
    return config


def read_excel_file(path, sheet_names):
    workbook = openpyxl.load_workbook(path, read_only=True)
    sheet1 = workbook[sheet_names.get('Setup_Sheet_Name')]
    sheet2 = workbook[sheet_names.get('Details_Sheet_Name')]
    sheet3 = workbook[sheet_names.get('DAQ_Sheet_Name')]

    return [sheet1, sheet2, sheet3]


def convert_into_json(sheets):
    sheet1, sheet2, sheet3 = sheets
    output_dict = {
        "test_setup": {},
        "test_details": {},
        "daq": {}
    }
    # Getting sheet 1 data
    for row in sheet1.iter_rows(min_row=2, max_row=2, values_only=True):
        test_setup = {
            "title": row[0],
            "author": row[1],
            "datetime": row[2],
        }

    output_dict['test_setup'] = test_setup
    test_details_keys = ["operator", "project_number", "part_number", "serial_number", "fluid", "fill_ratio",
                         "chiller_temp", "orientation", "test_destination", "clamp_pressure", "tim",
                         "comments", "timestamp"]
    test_details = dict.fromkeys(test_details_keys, None)

    # Getting sheet 2 data
    for row in sheet2.iter_rows(min_row=2, max_row=2, values_only=True):
        for i in range(0, len(test_details_keys) - 1):
            test_details[test_details_keys[i]] = round(row[i + i], 3) if isinstance(row[i + i], float) else row[i + i]
        if row[30] is not None and isinstance(row[30], datetime):

            test_details[test_details_keys[12]] = row[30].strftime("%m/%d/%Y %H:%M:%S")
        else:
            test_details[test_details_keys[12]] = row[30]

    output_dict['test_details'] = test_details

    # Getting sheet 3 data
    daq_headers = ["timestamp"]
    for row in sheet3.iter_rows(min_row=1, max_row=1, values_only=True):
        for i in range(1, len(row), 2):
            daq_headers.append(row[i])

    daq = []
    for row in sheet3.iter_rows(min_row=2, values_only=True):
        timestamp = row[0]
        if timestamp is not None and isinstance(timestamp, datetime):
            timestamp = timestamp.strftime("%m/%d/%Y %H:%M:%S %p")
        daq_data = [timestamp]
        for i in range(1, len(row), 2):
            daq_data.append(round(row[i], 3) if isinstance(row[i], float) else row[i])

        daq.append(daq_data)

    daq = {
        "headers": daq_headers,
        "data": daq
    }

    output_dict['daq'] = daq

    return output_dict


def write_to_file(output, output_folder_path):
    with open(output_folder_path + "/output.json", 'w', encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)


try:
    print("[INFO] Reading config file")
    config = read_config()
    paths = config['paths']
    sheet_names = config['sheet_names']
    excel_file_path = paths.get('excel_file_path')
    output_folder_path = paths.get('output_folder_path')
    print("[INFO] Validating paths")
    validate_path(excel_file_path)
    validate_path(output_folder_path)
    print("[INFO] Reading excel file")
    sheets = read_excel_file(excel_file_path, sheet_names)
    print("[INFO] Processing excel file")
    output = convert_into_json(sheets)
    print("[INFO] Writing to file")
    write_to_file(output, output_folder_path)


except KeyError as e:
    print("[ERROR] Key error: " + str(e))
    print("[INFO] Stopping script due to above error")
    exit(1)


except Exception as e:
    print("[ERROR] " + str(e))
    print("[INFO] Stopping script due to above error")
    exit(1)
